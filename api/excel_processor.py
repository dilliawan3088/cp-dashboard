"""
Excel file parser to extract data from uploaded files.
Handles multiple sheets and data formats.
"""
import pandas as pd
from openpyxl import load_workbook
from typing import List, Dict, Optional
import os
import json


class ExcelProcessor:
    """Process Excel files and extract poultry processing data"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.data_rows = []
        
    def load_file(self):
        """Load the Excel file"""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"File not found: {self.file_path}")
        
        self.workbook = load_workbook(self.file_path, data_only=True)
        return self
    
    def find_data_start_row(self, sheet, header_row: int = 1) -> Optional[int]:
        """Find the row where data starts (after headers)
        Row 1 is the header row, data starts from row 2
        """
        # Look for header row with expected columns
        expected_headers = ["NO", "TRUCK NO", "D/O Number", "FARM", "D/O Quantity", 
                           "BIRD COUNTER", "TOTAL SLAUGHTER", "DOA", "NON HALAL"]
        
        # Check if header row matches (row 1, columns A-I)
        header_values = []
        for col in range(1, 10):  # Columns A-I
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value:
                header_values.append(str(cell_value).strip().upper())
        
        # Debug: Print header row
        print(f"DEBUG: Header row {header_row} values: {header_values}")
        
        # Check if we found matching headers
        found_headers = any(h in " ".join(header_values) for h in expected_headers[:3])
        
        if found_headers:
            print(f"DEBUG: Found headers at row {header_row}, data starts at row {header_row + 1}")
            return header_row + 1  # Data starts after header row (row 2)
        
        # Try to find header row by searching
        for row in range(1, 10):
            row_values = []
            for col in range(1, 10):  # Columns A-I
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    row_values.append(str(cell_value).strip().upper())
            
            row_text = " ".join(row_values)
            if any(h in row_text for h in ["NO", "TRUCK", "D/O"]):
                print(f"DEBUG: Found headers at row {row}, data starts at row {row + 1}")
                return row + 1
        
        return None
    
    def extract_data_from_sheet(self, sheet_name: Optional[str] = None) -> List[Dict]:
        """Extract data from a specific sheet"""
        if not self.workbook:
            self.load_file()
        
        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        
        # Find data start row
        data_start_row = self.find_data_start_row(sheet)
        if not data_start_row:
            raise ValueError("Could not find data start row in Excel file")
        
        data_rows = []
        row_num = data_start_row
        
        # Extract data until we hit an empty row or "GRAND TOTAL"
        grand_total_row = None
        while True:
            # Check if this is a GRAND TOTAL row
            # Check multiple columns for "GRAND TOTAL" text
            is_grand_total = False
            for col in range(1, 10):  # Check columns A-I
                cell_value = sheet.cell(row=row_num, column=col).value
                if cell_value and "GRAND TOTAL" in str(cell_value).upper():
                    is_grand_total = True
                    break
            
            if is_grand_total:
                print(f"DEBUG: Found GRAND TOTAL at row {row_num}")
                # Extract grand total row - only numeric fields and row_number
                # Columns: E=D/O Quantity, F=BIRD COUNTER, G=TOTAL SLAUGHTER, H=DOA, I=NON HALAL
                grand_total_row = {
                    'row_number': row_num,
                    'do_quantity': self._get_cell_value(sheet, row_num, 5, int),  # Column E
                    'bird_counter': self._get_cell_value(sheet, row_num, 6, int),  # Column F
                    'total_slaughter': self._get_cell_value(sheet, row_num, 7, int),  # Column G
                    'doa': self._get_cell_value(sheet, row_num, 8, int),  # Column H
                    'non_halal': self._get_cell_value(sheet, row_num, 9, int),  # Column I
                }
                print(f"DEBUG: Stopping at row {row_num} - found GRAND TOTAL")
                break
            
            # Extract raw cell values for debugging
            raw_cells = []
            for col in range(1, 10):  # Columns A-I
                raw_value = sheet.cell(row=row_num, column=col).value
                raw_cells.append(raw_value)
            
            # Extract row data
            # Row 1 is header, data starts from row 2
            # Columns: A=NO, B=TRUCK NO, C=D/O Number, D=FARM, E=D/O Quantity, F=BIRD COUNTER, G=TOTAL SLAUGHTER, H=DOA, I=NON HALAL
            row_data = {
                'row_number': row_num,
                'no': self._get_cell_value(sheet, row_num, 1, int),  # Column A
                'truck_no': self._get_cell_value(sheet, row_num, 2, str),  # Column B
                'do_number': self._get_cell_value(sheet, row_num, 3, str),  # Column C
                'farm': self._get_cell_value(sheet, row_num, 4, str),  # Column D
                'do_quantity': self._get_cell_value(sheet, row_num, 5, int),  # Column E
                'bird_counter': self._get_cell_value(sheet, row_num, 6, int),  # Column F
                'total_slaughter': self._get_cell_value(sheet, row_num, 7, int),  # Column G
                'doa': self._get_cell_value(sheet, row_num, 8, int),  # Column H
                'non_halal': self._get_cell_value(sheet, row_num, 9, int),  # Column I
            }
            
            # Debug: Print first few rows
            if len(data_rows) < 3:
                print(f"DEBUG: Row {row_num} raw cells: {raw_cells}")
                print(f"DEBUG: Row {row_num} extracted: {row_data}")
            
            # Check if row has any data
            if any(row_data.values()):
                data_rows.append(row_data)
            else:
                # Empty row, might be end of data
                if len(data_rows) > 0:
                    print(f"DEBUG: Stopping at row {row_num} - empty row after data")
                    break
            
            row_num += 1
            
            # Safety limit
            if row_num > 1000:
                print(f"DEBUG: Stopping at row {row_num} - safety limit reached")
                break
        
        print(f"DEBUG: Extracted {len(data_rows)} data rows")
        return data_rows, grand_total_row
    
    def _get_cell_value(self, sheet, row: int, col: int, data_type):
        """Get cell value and convert to specified type"""
        try:
            value = sheet.cell(row=row, column=col).value
            
            # Debug: Show conversion for first few cells
            if row <= 5 and col <= 5:
                print(f"DEBUG: Cell({row},{col}) raw={value} type={type(value)} target={data_type}")
            
            if value is None:
                return None
            
            # Handle string values
            if data_type == str:
                result = str(value).strip() if value else None
                return result
            
            # Handle numeric values
            if data_type == int:
                if isinstance(value, (int, float)):
                    return int(value)
                elif isinstance(value, str):
                    # Try to extract number from string
                    cleaned = value.replace(',', '').replace(' ', '').strip()
                    if cleaned:
                        try:
                            return int(float(cleaned))
                        except (ValueError, TypeError):
                            print(f"DEBUG: Could not convert '{value}' to int")
                            return None
                return None
            
            return value
        except (ValueError, TypeError) as e:
            print(f"DEBUG: Error converting cell({row},{col}): {e}")
            return None
    
    def extract_all_data(self) -> tuple:
        """Extract data from all sheets or active sheet
        
        Returns:
            Tuple of (data_rows, grand_total_row)
        """
        if not self.workbook:
            self.load_file()
        
        all_data = []
        grand_total = None
        
        # Try to extract from active sheet first
        try:
            data, grand_total_row = self.extract_data_from_sheet()
            all_data.extend(data)
            if grand_total_row:
                grand_total = grand_total_row
        except Exception as e:
            print(f"Error extracting from active sheet: {e}")
        
        # If no data found, try all sheets
        if not all_data:
            for sheet_name in self.workbook.sheetnames:
                try:
                    data, grand_total_row = self.extract_data_from_sheet(sheet_name)
                    if data:
                        all_data.extend(data)
                        if grand_total_row:
                            grand_total = grand_total_row
                        break  # Use first sheet with data
                except Exception as e:
                    print(f"Error extracting from sheet {sheet_name}: {e}")
                    continue
        
        return all_data, grand_total
    
    def get_detail_analysis(self, sheet_name: Optional[str] = None) -> List[Dict]:
        """Extract detail analysis table from columns C, D, L, M, N (rows 2-14)"""
        if not self.workbook:
            self.load_file()
        
        sheet = sheet_name if sheet_name else self.workbook.active
        
        if isinstance(sheet, str):
            sheet = self.workbook[sheet]
        
        detail_analysis = []
        
        # Get column headers from row 2
        header_c2 = self._get_cell_value(sheet, 2, 3, str) or ""  # Column C, Row 2
        header_d2 = self._get_cell_value(sheet, 2, 4, str) or ""  # Column D, Row 2
        header_l2 = self._get_cell_value(sheet, 2, 12, str) or ""  # Column L, Row 2
        header_m2 = self._get_cell_value(sheet, 2, 13, str) or ""  # Column M, Row 2
        header_n2 = self._get_cell_value(sheet, 2, 14, str) or ""  # Column N, Row 2
        
        print(f"DEBUG: Detail analysis headers - C2: {header_c2}, D2: {header_d2}, L2: {header_l2}, M2: {header_m2}, N2: {header_n2}")
        
        # Extract data from rows 3-14
        for row in range(3, 15):  # Rows 3 to 14
            row_data = {
                header_c2: self._get_cell_value(sheet, row, 3, str),  # Column C
                header_d2: self._get_cell_value(sheet, row, 4, str),  # Column D
                header_l2: self._get_cell_value(sheet, row, 12, int),  # Column L
                header_m2: self._get_cell_value(sheet, row, 13, int),  # Column M
                header_n2: self._get_cell_value(sheet, row, 14, int),  # Column N
            }
            
            # Only add if row has some data
            if any(row_data.values()):
                detail_analysis.append(row_data)
        
        print(f"DEBUG: Extracted {len(detail_analysis)} detail analysis rows")
        return detail_analysis
    
    def get_summary_data(self, sheet_name: Optional[str] = None) -> Dict:
        """Extract summary data from specific cells: L23/M23/N23 for Broiler, L24/M24/N24 for Breeder"""
        if not self.workbook:
            self.load_file()
        
        sheet = sheet_name if sheet_name else self.workbook.active
        
        if isinstance(sheet, str):
            sheet = self.workbook[sheet]
        
        # Extract from specific cells
        # Broiler: Row 23, Columns L(12), M(13), N(14)
        # Breeder: Row 24, Columns L(12), M(13), N(14)
        summary = {
            'broiler': {
                'counter_plus_doa': self._get_cell_value(sheet, 23, 12, int) or 0,  # L23
                'do_quantity': self._get_cell_value(sheet, 23, 13, int) or 0,  # M23
                'different': self._get_cell_value(sheet, 23, 14, int) or 0,  # N23
            },
            'breeder': {
                'counter_plus_doa': self._get_cell_value(sheet, 24, 12, int) or 0,  # L24
                'do_quantity': self._get_cell_value(sheet, 24, 13, int) or 0,  # M24
                'different': self._get_cell_value(sheet, 24, 14, int) or 0,  # N24
            }
        }
        
        print(f"DEBUG: Summary data - Broiler: {summary['broiler']}, Breeder: {summary['breeder']}")
        
        return summary


def process_excel_file(file_path: str) -> Dict:
    """
    Main function to process an Excel file and return extracted data.
    Only handles the first/main data table.
    
    Returns:
        Dictionary with 'raw_data' and 'grand_total' keys
    """
    processor = ExcelProcessor(file_path)
    processor.load_file()
    
    raw_data, grand_total = processor.extract_all_data()
    
    result = {
        'raw_data': raw_data,
        'grand_total': grand_total
    }
    
    # JSON files will be saved in main.py to avoid duplication
    return result
